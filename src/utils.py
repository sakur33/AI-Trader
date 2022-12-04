import openpyxl
from datetime import datetime, date

############### EXCEL ####################


def candles_to_excel(candles, path, f_name, exec_start):
    exec_start = get_time()
    if candles == False:
        print("Error: No Candles!")
        # error
        return False
    try:
        wb = openpyxl.Workbook()
        wspace = wb.active
        wspace.title = "Candles"
        for pages in candles:
            wspace.append(list(pages.values()))
        wb.save(path + f_name)
        # success
        return True
    except:
        # error
        return False


def get_candles_from_excel(address, name):
    temp1 = []
    wb = openpyxl.load_workbook(address + name)
    wsp = wb.active
    for rows in wb.active.iter_rows(min_row=0, max_row=1000000):
        temp = {}
        i = 0
        for cell in rows:
            if i == 0 and cell.value == None:
                return temp1
            elif i == 0 and cell.value != None:
                temp["datetime"] = cell.value
            elif i == 1 and cell.value != None:
                temp["open"] = cell.value
            elif i == 2 and cell.value != None:
                temp["close"] = cell.value
            elif i == 3 and cell.value != None:
                temp["high"] = cell.value
            elif i == 4 and cell.value != None:
                temp["low"] = cell.value
            i += 1
        temp1.append(temp)


def get_time():
    time = date.today().strftime("%m/%d/%Y %H:%M:%S%f")
    time = datetime.strptime(time, "%m/%d/%Y %H:%M:%S%f")
    return time


def to_milliseconds(days=0, hours=0, minutes=0):
    milliseconds = (
        (days * 24 * 60 * 60 * 1000) + (hours * 60 * 60 * 1000) + (minutes * 60 * 1000)
    )
    return milliseconds


def time_conversion(date):
    start = "01/01/1970 00:00:00"
    start = datetime.strptime(start, "%m/%d/%Y %H:%M:%S")
    date = datetime.strptime(date, "%m/%d/%Y %H:%M:%S")
    final_date = date - start
    temp = str(final_date)
    temp1, temp2 = temp.split(", ")
    hours, minutes, seconds = temp2.split(":")
    days = final_date.days
    days = int(days)
    hours = int(hours)
    hours += 2
    minutes = int(minutes)
    seconds = int(seconds)
    time = (
        (days * 24 * 60 * 60 * 1000)
        + (hours * 60 * 60 * 1000)
        + (minutes * 60 * 1000)
        + (seconds * 1000)
    )
    return time






